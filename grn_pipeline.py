#!/usr/bin/env python3
"""
Gene Regulatory Network Construction and Analysis Pipeline
for Streptococcus pneumoniae genetic interaction data.

Integrates:
  - Dual Tn-seq genetic interactions (table_S1 + table_S3)
  - Sup-seq genetic suppression edges (Supplementary Data 1)
  - STRING PPI edges (API, taxon 1313)
  - RNA-seq differential expression (Supplementary Data 6)
  - RB Tn-seq single-mutant fitness (table_S4)

Produces:
  - Network topology analysis (centrality, hubs)
  - Community detection (Louvain)
  - Perturbed pathway identification (DE enrichment + COG characterization)
  - Static SVG figures
  - Interactive PyVis HTML
  - Community-facing webpage (index.html + data.js)
"""

import argparse
import json
import os
import re
import sys
import time
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Set, Tuple

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib import rcParams as mpl_rc
import networkx as nx
import numpy as np
import openpyxl
import pandas as pd
import requests
import seaborn as sns
from scipy import stats

# ── Matplotlib defaults ──
mpl_rc['font.family'] = ['Liberation Sans', 'Arimo', 'DejaVu Sans']
mpl_rc['svg.fonttype'] = 'none'
sns.set_style('whitegrid')

# ── Constants ──
QUERY_GENES = ['ackA', 'cdsA', 'yqeH', 'ccrZ', 'ezrA', 'tsaC', 'mreC']
# Fallback locus tags for query genes not in RNA-seq annotation
# (sourced from PneumoWiki — S. pneumoniae D39, CP000410.2)
QUERY_GENE_LOCI = {
    'ackA': 'SPD_1853', 'cdsA': 'SPD_0244', 'yqeH': 'SPD_1559',
    'ccrZ': 'SPD_0476', 'ezrA': 'SPD_0710', 'tsaC': 'SPD_0908',
    'mreC': 'SPD_2045',
}
# COG functional category one-letter codes mapped from bracket-text keywords
COG_CATEGORY_MAP = [
    ('Translation', 'J'), ('Transcription', 'K'),
    ('DNA replication', 'L'), ('Cell division', 'D'),
    ('Posttranslational modification', 'O'), ('Cell envelope', 'M'),
    ('Cell motility', 'N'), ('Inorganic ion transport', 'P'),
    ('Signal transduction', 'T'), ('Energy production', 'C'),
    ('Carbohydrate transport', 'G'), ('Amino acid transport', 'E'),
    ('Nucleotide transport', 'F'), ('Coenzyme metabolism', 'H'),
    ('Lipid metabolism', 'I'), ('Secondary metabolites', 'Q'),
    ('General function', 'R'), ('Function unknown', 'S'),
    ('Defense mechanisms', 'V'), ('Extracellular structures', 'W'),
    ('Intracellular trafficking', 'U'),
]
COG_CODE_NAMES = {
    'J': 'Translation', 'K': 'Transcription', 'L': 'DNA replication/repair',
    'D': 'Cell division', 'O': 'Posttranslational modification',
    'M': 'Cell envelope', 'N': 'Cell motility', 'P': 'Inorganic ion transport',
    'T': 'Signal transduction', 'C': 'Energy production',
    'G': 'Carbohydrate metabolism', 'E': 'Amino acid metabolism',
    'F': 'Nucleotide metabolism', 'H': 'Coenzyme metabolism',
    'I': 'Lipid metabolism', 'Q': 'Secondary metabolites',
    'R': 'General function', 'S': 'Function unknown',
    'V': 'Defense mechanisms', 'W': 'Extracellular structures',
    'U': 'Intracellular trafficking',
}
STRING_API = 'https://string-db.org/api'
SPECIES = 1313  # S. pneumoniae


def extract_cog_category(cog_str: str) -> str:
    """Extract COG one-letter functional category from full COG description.
    e.g. 'gnl|CDD|...;COG1197, Mfd, ... [DNA replication.../ Transcription].'
    -> 'L'"""
    if not cog_str or cog_str == '-':
        return '-'
    # Extract text in brackets [...]
    m = re.search(r'\[([^\]]+)\]', cog_str)
    if not m:
        return '-'
    bracket_text = m.group(1)
    # Map first matching keyword to one-letter code
    for keyword, code in COG_CATEGORY_MAP:
        if keyword.lower() in bracket_text.lower():
            return code
    return '-'


@dataclass
class Config:
    table_s3: str = ''
    table_s1: str = ''
    sup_data1: str = ''
    sup_data6: str = ''
    table_s4: str = ''
    output_dir: str = '/mnt/results'
    z_threshold: float = 4.0
    string_score: float = 0.7
    string_species: int = SPECIES
    de_fdr: float = 0.05
    de_logfc: float = 1.0  # effect-size threshold for DE calling
    top_hubs: int = 20
    data_dir: str = ''


# ═══════════════════════════════════════════════════════════════════════
# STEP 1: Data Loading and Parsing
# ═══════════════════════════════════════════════════════════════════════

def load_dual_tnseq_s3(path: str) -> pd.DataFrame:
    """Load significant hits from table_S3 (header at row 12, 1-indexed)."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb['confident_hit_table']
    rows = list(ws.iter_rows(min_row=12, max_row=ws.max_row, values_only=True))
    wb.close()
    df = pd.DataFrame(rows[1:], columns=rows[0])
    df = df.dropna(subset=['locusId1', 'locusId2'])
    df['locusId1'] = df['locusId1'].astype(str).str.strip()
    df['locusId2'] = df['locusId2'].astype(str).str.strip()
    # Parse zStrains and readRatio
    for col in ['zStrains(10-90%ORF)', 'readRatio(10-90%ORF)',
                'zStrains(0-100%ORF)', 'readRatio(0-100%ORF)']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    df['source'] = 'Dual Tn-seq (S3)'
    print(f"  table_S3: {len(df)} significant hits loaded")
    return df


def load_dual_tnseq_s1(path: str, s3_pairs: Set[Tuple[str, str]],
                       z_threshold: float) -> pd.DataFrame:
    """Load additional edges from table_S1 with |zStrains| > threshold."""
    wb = openpyxl.load_workbook(path, read_only=True)
    all_rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Header at row 16 (1-indexed), data from row 17
        header = None
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i < 16:
                continue
            if i == 16:
                header = row
                continue
            # Filter: |zStrains(10-90%)| > threshold and not in S3
            try:
                z_val = row[13]  # zStrains(10-90%) column (0-indexed: 13)
                if z_val is None:
                    continue
                z_val = float(z_val)
            except (ValueError, TypeError):
                continue
            if abs(z_val) <= z_threshold:
                continue
            locus1 = str(row[1]).strip() if row[1] else ''
            locus2 = str(row[2]).strip() if row[2] else ''
            if not locus1 or not locus2:
                continue
            pair = tuple(sorted([locus1, locus2]))
            if pair in s3_pairs:
                continue
            all_rows.append(row)
        print(f"  table_S1 sheet '{sheet_name}': scanned, {len(all_rows)} cumulative extra edges")
    wb.close()

    if not all_rows:
        return pd.DataFrame()
    df = pd.DataFrame(all_rows, columns=header)
    df = df.dropna(subset=['locusId1', 'locusId2'])
    df['locusId1'] = df['locusId1'].astype(str).str.strip()
    df['locusId2'] = df['locusId2'].astype(str).str.strip()
    for col in ['zStrains(10-90%)', 'readRatio(10-90%)',
                'zStrains(0-100%)', 'readRatio(0-100%)']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    df['source'] = 'Dual Tn-seq (S1 strict)'
    print(f"  table_S1: {len(df)} additional edges (|z| > {z_threshold})")
    return df


def load_supseq_hits(path: str, locus_map: Dict[str, str]) -> pd.DataFrame:
    """Load Sup-seq suppression hits from Supplementary Data 1."""
    wb = openpyxl.load_workbook(path, read_only=True)
    all_hits = []
    for sheet_name in wb.sheetnames:
        if sheet_name in ('Legend and notes', 'Summary'):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        header = rows[1]  # Row 2 (0-indexed: 1) is the header
        for row in rows[2:]:
            if row[0] is None:
                continue
            hit_locus = str(row[0]).strip() if row[0] else ''
            # Skip non-data rows (summary headers, footers, repeated column headers)
            if not re.match(r'SPD_\d', hit_locus):
                continue
            gene_names = str(row[1]).strip() if row[1] else ''
            gene_dir = str(row[2]).strip() if row[2] else ''
            log10_rr_pos = row[3]
            log10_sr_pos = row[4]
            log10_rr_neg = row[5]
            log10_sr_neg = row[6]
            polarity = str(row[10]).strip() if len(row) > 10 and row[10] else ''

            # Expand locus ranges
            expanded = expand_locus_range(hit_locus, locus_map)
            for locus in expanded:
                all_hits.append({
                    'query_gene': sheet_name,
                    'hit_locus': locus,
                    'gene_names': gene_names,
                    'gene_direction': gene_dir,
                    'log10_rr_pos': log10_rr_pos,
                    'log10_sr_pos': log10_sr_pos,
                    'log10_rr_neg': log10_rr_neg,
                    'log10_sr_neg': log10_sr_neg,
                    'polarity': polarity,
                    'source': 'Sup-seq',
                })
    wb.close()
    df = pd.DataFrame(all_hits)
    print(f"  Sup Data 1: {len(df)} suppression hits across {df['query_gene'].nunique()} query genes")
    return df


def expand_locus_range(locus_str: str, locus_map: Dict[str, str]) -> List[str]:
    """Expand locus ranges like 'SPD_0033-SPD_0035' to individual loci."""
    locus_str = locus_str.strip()
    # Handle comma-separated
    if ',' in locus_str:
        parts = locus_str.split(',')
        result = []
        for p in parts:
            result.extend(expand_locus_range(p, locus_map))
        return result
    # Handle range
    m = re.match(r'(SPD_\d+)-(SPD_\d+)', locus_str)
    if m:
        start = int(m.group(1).replace('SPD_', ''))
        end = int(m.group(2).replace('SPD_', ''))
        result = []
        for i in range(start, end + 1):
            tag = f'SPD_{i:04d}'
            if tag in locus_map:
                result.append(tag)
        return result if result else [locus_str.split('-')[0]]
    # Single locus
    return [locus_str]


def load_rnaseq_de(path: str) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Load RNA-seq DE + annotations from Sup Data 6."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb['log2FC & FDR']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    df = pd.DataFrame(rows[1:], columns=rows[0])
    df = df.dropna(subset=['GeneID'])
    # Clean locus tags
    df['locus_tag'] = df['locus_tag'].astype(str).str.strip()
    df['logFC'] = pd.to_numeric(df['logFC'], errors='coerce')
    df['FDR'] = pd.to_numeric(df['FDR'], errors='coerce')
    # Build locus -> gene name mapping
    locus_map = {}
    for _, row in df.iterrows():
        lt = str(row.get('locus_tag', '')).strip()
        gn = str(row.get('GeneName', '')).strip()
        if lt and gn and gn != lt:
            locus_map[lt] = gn
    print(f"  Sup Data 6: {len(df)} genes with DE data, {len(locus_map)} named genes")
    return df, locus_map


def load_rbtnseq_fitness(path: str) -> pd.DataFrame:
    """Load RB Tn-seq fitness from table_S4."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb['Fitness(median) values']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    # Row 8 (1-indexed) = index 7: strain names
    # Row 12 (1-indexed) = index 11: header 'locusID (below)'
    # Data from row 13 (1-indexed) = index 12
    header_row = rows[11]  # 'locusID (below)', '', '', ...
    data_rows = rows[12:]
    fitness_data = []
    for row in data_rows:
        if row[0] is None:
            continue
        locus = str(row[0]).strip()
        values = []
        for v in row[1:]:
            if v is not None and v != 'N/A' and v != '':
                try:
                    values.append(float(v))
                except (ValueError, TypeError):
                    pass
        if values:
            fitness_data.append({
                'locusId': locus,
                'fitness_median': float(np.median(values)),
                'fitness_std': float(np.std(values)) if len(values) > 1 else 0.0,
            })
    df = pd.DataFrame(fitness_data)
    print(f"  table_S4: {len(df)} genes with fitness data")
    return df


# ═══════════════════════════════════════════════════════════════════════
# STEP 2: STRING PPI Retrieval
# ═══════════════════════════════════════════════════════════════════════

def fetch_string_ppi(gene_names: List[str], species: int,
                     score_threshold: float) -> pd.DataFrame:
    """Query STRING API for PPI edges — all identifiers in one request to
    capture cross-batch interactions. Falls back to overlapping batches if
    the API enforces a per-request identifier limit."""
    all_edges = []
    score_param = int(score_threshold * 1000)

    def _query_network(identifiers: List[str], label: str) -> int:
        """Query STRING /network endpoint and append edges. Returns count."""
        params = {
            'identifiers': '\n'.join(identifiers),
            'species': species,
            'required_score': score_param,
            'limit': 10000,
        }
        try:
            resp = requests.get(
                f'{STRING_API}/tsv-no-header/network',
                params=params, timeout=120
            )
            if resp.status_code == 200 and resp.text.strip():
                count = 0
                for line in resp.text.strip().split('\n'):
                    parts = line.split('\t')
                    if len(parts) >= 7:
                        all_edges.append({
                            'stringId_A': parts[0],
                            'stringId_B': parts[1],
                            'gene_a': parts[2],
                            'gene_b': parts[3],
                            'score': float(parts[5]),
                            'source': 'STRING',
                        })
                        count += 1
                print(f"    {label}: +{count} edges")
                return count
            else:
                print(f"    {label}: HTTP {resp.status_code}, {len(resp.text)} bytes")
                return 0
        except Exception as e:
            print(f"    {label}: ERROR - {e}")
            return 0

    print(f"  STRING API: {len(gene_names)} genes (score > {score_threshold})")

    # Try all identifiers in one request first
    n = _query_network(gene_names, 'All-at-once')
    time.sleep(0.5)

    if n == 0:
        # Fallback: batched queries with cross-batch coverage
        # Use overlapping batches so cross-batch interactions are captured
        batch_size = 450
        batches = [gene_names[i:i+batch_size]
                   for i in range(0, len(gene_names), batch_size)]
        print(f"  Fallback: {len(batches)} batches with cross-batch queries")
        for i, batch in enumerate(batches):
            _query_network(batch, f'Batch {i+1}/{len(batches)}')
            time.sleep(0.5)
        # Cross-batch: query each batch paired with the next
        for i in range(len(batches) - 1):
            combined = batches[i] + batches[i + 1]
            _query_network(combined, f'Cross-batch {i+1}-{i+2}')
            time.sleep(0.5)

    df = pd.DataFrame(all_edges)
    if not df.empty:
        df = df.drop_duplicates(subset=['gene_a', 'gene_b'])
    print(f"  STRING API: {len(df)} PPI edges retrieved")
    return df


# ═══════════════════════════════════════════════════════════════════════
# STEP 3: Network Construction
# ═══════════════════════════════════════════════════════════════════════

def build_network(s3_df: pd.DataFrame, s1_df: pd.DataFrame,
                  supseq_df: pd.DataFrame, string_df: pd.DataFrame,
                  rnaseq_df: pd.DataFrame, fitness_df: pd.DataFrame,
                  locus_map: Dict[str, str], config: Config) -> nx.MultiDiGraph:
    """Build multi-layer network."""
    G = nx.MultiDiGraph()

    # ── Collect all nodes ──
    all_loci = set()
    # From S3
    all_loci.update(s3_df['locusId1'].tolist())
    all_loci.update(s3_df['locusId2'].tolist())
    # From S1
    if not s1_df.empty:
        all_loci.update(s1_df['locusId1'].tolist())
        all_loci.update(s1_df['locusId2'].tolist())
    # From Sup-seq
    if not supseq_df.empty:
        all_loci.update(supseq_df['hit_locus'].tolist())
        # Query genes -> map to locus tags
        for qg in supseq_df['query_gene'].unique():
            qlocus = find_locus_for_gene(qg, locus_map, rnaseq_df)
            if qlocus:
                all_loci.add(qlocus)
    # From STRING
    if not string_df.empty:
        for _, row in string_df.iterrows():
            la = gene_to_locus(row['gene_a'], locus_map, rnaseq_df)
            lb = gene_to_locus(row['gene_b'], locus_map, rnaseq_df)
            if la: all_loci.add(la)
            if lb: all_loci.add(lb)

    # ── Add nodes with attributes ──
    rnaseq_indexed = rnaseq_df.set_index('locus_tag') if not rnaseq_df.empty else pd.DataFrame()
    fitness_indexed = fitness_df.set_index('locusId') if not fitness_df.empty else pd.DataFrame()

    for locus in sorted(all_loci):
        attrs = {
            'locus_tag': locus,
            'gene_name': locus_map.get(locus, locus),
            'logFC': None, 'FDR': None, 'is_de': False,
            'cog_category': '-', 'cog_description': '-',
            'go_bp': '-', 'kegg': '-', 'product': '-',
            'fitness_median': None, 'is_essential': False,
        }
        # RNA-seq attributes
        if not rnaseq_indexed.empty and locus in rnaseq_indexed.index:
            row = rnaseq_indexed.loc[locus]
            attrs['logFC'] = float(row['logFC']) if pd.notna(row['logFC']) else None
            attrs['FDR'] = float(row['FDR']) if pd.notna(row['FDR']) else None
            # DE requires both FDR and effect-size thresholds
            attrs['is_de'] = bool(
                attrs['FDR'] is not None and attrs['FDR'] < config.de_fdr
                and attrs['logFC'] is not None and abs(attrs['logFC']) >= config.de_logfc
            )
            cog_full = str(row.get('COG', '-'))
            attrs['cog_category'] = extract_cog_category(cog_full) if cog_full else '-'
            attrs['cog_description'] = cog_full if cog_full else '-'
            attrs['product'] = str(row.get('Product', '-'))
            attrs['go_bp'] = str(row.get('GO_Biological_Process', '-'))
            attrs['kegg'] = str(row.get('KEGG', '-'))
        # Fitness
        if not fitness_indexed.empty and locus in fitness_indexed.index:
            attrs['fitness_median'] = float(fitness_indexed.loc[locus, 'fitness_median'])
        # Essential
        gene_name = locus_map.get(locus, '')
        if gene_name in QUERY_GENES:
            attrs['is_essential'] = True
        G.add_node(locus, **attrs)

    # ── Add edges: Dual Tn-seq (S3) ──
    for _, row in s3_df.iterrows():
        z = row.get('zStrains(10-90%ORF)')
        if pd.isna(z):
            z = row.get('zStrains(0-100%ORF)')
        G.add_edge(row['locusId1'], row['locusId2'],
                   layer='genetic_interaction',
                   source='Dual Tn-seq',
                   subtype=row.get('threshold', 'medium-only'),
                   zscore=float(z) if pd.notna(z) else 0.0,
                   read_ratio=float(row.get('readRatio(10-90%ORF)', 0)) if pd.notna(row.get('readRatio(10-90%ORF)')) else 0.0,
                   interaction_class='negative' if (pd.notna(z) and z < 0) else 'positive',
                   validation=str(row.get('validation', 'Untested')))

    # ── Add edges: Dual Tn-seq (S1 strict) ──
    if not s1_df.empty:
        for _, row in s1_df.iterrows():
            z = row.get('zStrains(10-90%)')
            if pd.isna(z):
                z = row.get('zStrains(0-100%)')
            G.add_edge(row['locusId1'], row['locusId2'],
                       layer='genetic_interaction',
                       source='Dual Tn-seq (strict)',
                       subtype='strict',
                       zscore=float(z) if pd.notna(z) else 0.0,
                       read_ratio=float(row.get('readRatio(10-90%)', 0)) if pd.notna(row.get('readRatio(10-90%)')) else 0.0,
                       interaction_class='negative' if (pd.notna(z) and z < 0) else 'positive',
                       validation='Untested')

    # ── Add edges: Sup-seq (directed) ──
    if not supseq_df.empty:
        for _, row in supseq_df.iterrows():
            qlocus = find_locus_for_gene(row['query_gene'], locus_map, rnaseq_df)
            if not qlocus:
                continue
            G.add_edge(row['hit_locus'], qlocus,
                       layer='suppression',
                       source='Sup-seq',
                       subtype='suppression',
                       query_gene=row['query_gene'],
                       polarity=row.get('polarity', ''),
                       gene_direction=row.get('gene_direction', ''),
                       interaction_class='suppression',
                       validation='Published')

    # ── Add edges: STRING PPI ──
    if not string_df.empty:
        for _, row in string_df.iterrows():
            la = gene_to_locus(row['gene_a'], locus_map, rnaseq_df)
            lb = gene_to_locus(row['gene_b'], locus_map, rnaseq_df)
            if la and lb and la != lb:
                G.add_edge(la, lb,
                           layer='ppi',
                           source='STRING',
                           subtype='PPI',
                           score=float(row['score']),
                           interaction_class='PPI',
                           validation='Computational')

    # ── Print stats ──
    n_gi = sum(1 for _, _, d in G.edges(data=True) if d.get('layer') == 'genetic_interaction')
    n_sup = sum(1 for _, _, d in G.edges(data=True) if d.get('layer') == 'suppression')
    n_ppi = sum(1 for _, _, d in G.edges(data=True) if d.get('layer') == 'ppi')
    print(f"  Network: {G.number_of_nodes()} nodes, {G.number_of_edges()} edges")
    print(f"    Genetic interactions: {n_gi}")
    print(f"    Suppression edges: {n_sup}")
    print(f"    PPI edges: {n_ppi}")
    return G


def find_locus_for_gene(gene_name: str, locus_map: Dict[str, str],
                        rnaseq_df: pd.DataFrame) -> Optional[str]:
    """Find locus tag for a gene name."""
    # Reverse mapping from RNA-seq annotation
    for locus, name in locus_map.items():
        if name == gene_name:
            return locus
    # Try direct lookup in rnaseq
    if not rnaseq_df.empty:
        matches = rnaseq_df[rnaseq_df['GeneName'] == gene_name]
        if not matches.empty:
            return str(matches.iloc[0]['locus_tag']).strip()
    # Fallback: hardcoded mapping for query genes not in RNA-seq annotation
    if gene_name in QUERY_GENE_LOCI:
        return QUERY_GENE_LOCI[gene_name]
    return None


def gene_to_locus(gene_name: str, locus_map: Dict[str, str],
                  rnaseq_df: pd.DataFrame) -> Optional[str]:
    """Map STRING gene name to SPD_ locus tag."""
    return find_locus_for_gene(gene_name, locus_map, rnaseq_df)


# ═══════════════════════════════════════════════════════════════════════
# STEP 4: Network Topology Analysis
# ═══════════════════════════════════════════════════════════════════════

def compute_centrality(G: nx.MultiDiGraph) -> pd.DataFrame:
    """Compute centrality metrics on undirected projection."""
    # Undirected simple graph
    Gu = nx.Graph()
    for u, v, d in G.edges(data=True):
        if Gu.has_edge(u, v):
            Gu[u][v]['weight'] = max(Gu[u][v].get('weight', 0), abs(d.get('zscore', d.get('score', 1))))
        else:
            Gu.add_edge(u, v, weight=abs(d.get('zscore', d.get('score', 1))))

    print("  Computing degree centrality...")
    deg = nx.degree_centrality(Gu)
    print("  Computing betweenness centrality...")
    betw = nx.betweenness_centrality(Gu, weight=None)
    print("  Computing closeness centrality...")
    close = nx.closeness_centrality(Gu)
    print("  Computing eigenvector centrality...")
    try:
        eigen = nx.eigenvector_centrality(Gu, max_iter=1000)
    except Exception:
        eigen = {n: 0 for n in Gu.nodes()}
    print("  Computing clustering coefficient...")
    clust = nx.clustering(Gu)

    rows = []
    for node in G.nodes():
        rows.append({
            'locus_tag': node,
            'gene_name': G.nodes[node].get('gene_name', node),
            'degree': deg.get(node, 0),
            'degree_raw': Gu.degree(node),
            'betweenness': betw.get(node, 0),
            'closeness': close.get(node, 0),
            'eigenvector': eigen.get(node, 0),
            'clustering': clust.get(node, 0),
            'is_essential': G.nodes[node].get('is_essential', False),
            'is_de': G.nodes[node].get('is_de', False),
            'logFC': G.nodes[node].get('logFC'),
            'FDR': G.nodes[node].get('FDR'),
            'cog_category': G.nodes[node].get('cog_category', '-'),
            'product': G.nodes[node].get('product', '-'),
        })
    df = pd.DataFrame(rows).sort_values('betweenness', ascending=False)
    print(f"  Centrality computed for {len(df)} nodes")
    return df, Gu


# ═══════════════════════════════════════════════════════════════════════
# STEP 5: Community Detection
# ═══════════════════════════════════════════════════════════════════════

def detect_communities(Gu: nx.Graph) -> Tuple[Dict[str, int], float]:
    """Run Louvain community detection."""
    try:
        communities = nx.community.louvain_communities(Gu, seed=42)
    except AttributeError:
        # Fallback: greedy modularity
        communities = nx.community.greedy_modularity_communities(Gu)

    comm_map = {}
    for i, comm in enumerate(communities):
        for node in comm:
            comm_map[node] = i
    modularity = nx.community.modularity(Gu, communities)
    print(f"  Communities: {len(communities)}, modularity: {modularity:.4f}")
    sizes = {}
    for c in comm_map.values():
        sizes[c] = sizes.get(c, 0) + 1
    print(f"  Community sizes: {dict(sorted(sizes.items(), key=lambda x: -x[1])[:10])}")
    return comm_map, modularity


# ═══════════════════════════════════════════════════════════════════════
# STEP 6: Perturbed Pathway Identification
# ═══════════════════════════════════════════════════════════════════════

def test_de_enrichment(comm_map: Dict[str, int], G: nx.MultiDiGraph,
                       de_fdr: float) -> pd.DataFrame:
    """Test each community for enrichment of differentially expressed genes."""
    # Get DE genes
    de_genes = set()
    total_genes = set()
    for node in G.nodes():
        total_genes.add(node)
        if G.nodes[node].get('is_de', False):
            de_genes.add(node)

    N = len(total_genes)
    K = len(de_genes)
    results = []
    communities = sorted(set(comm_map.values()))
    for c in communities:
        comm_genes = set(n for n, ci in comm_map.items() if ci == c)
        n = len(comm_genes)
        k = len(comm_genes & de_genes)
        # Fisher's exact test
        table = [[k, n - k], [K - k, N - K - (n - k)]]
        try:
            _, pval = stats.fisher_exact(table, alternative='greater')
        except Exception:
            pval = 1.0
        results.append({
            'community': c,
            'size': n,
            'de_count': k,
            'de_expected': n * K / N if N > 0 else 0,
            'de_enrichment': (k / n) / (K / N) if n > 0 and K > 0 else 0,
            'pvalue': pval,
        })
    df = pd.DataFrame(results)
    # BH-FDR correction
    from statsmodels.stats.multitest import multipletests
    if len(df) > 1:
        _, df['padj'], _, _ = multipletests(df['pvalue'], method='fdr_bh')
    else:
        df['padj'] = df['pvalue']
    df['is_perturbed'] = df['padj'] < 0.05
    n_perturbed = df['is_perturbed'].sum()
    print(f"  DE enrichment: {n_perturbed}/{len(df)} communities perturbed (padj < 0.05)")
    return df


def characterize_cog(comm_map: Dict[str, int], G: nx.MultiDiGraph) -> pd.DataFrame:
    """Characterize each community by dominant COG functional category (one-letter code)."""
    results = []
    communities = sorted(set(comm_map.values()))
    # Get genome-wide COG distribution
    all_cog = {}
    for node in G.nodes():
        cog = G.nodes[node].get('cog_category', '-')
        all_cog[cog] = all_cog.get(cog, 0) + 1
    N = sum(all_cog.values())

    for c in communities:
        comm_genes = [n for n, ci in comm_map.items() if ci == c]
        cog_counts = {}
        for g in comm_genes:
            cog = G.nodes[g].get('cog_category', '-')
            cog_counts[cog] = cog_counts.get(cog, 0) + 1
        # Dominant COG (excluding '-' and 'S' unknown function)
        dominant = max((k for k in cog_counts if k not in ('-', 'S')),
                       key=lambda k: cog_counts[k], default='-')
        dom_count = cog_counts.get(dominant, 0)
        n = len(comm_genes)
        # Enrichment test
        K = all_cog.get(dominant, 0)
        table = [[dom_count, n - dom_count], [K - dom_count, N - K - (n - dom_count)]]
        try:
            _, pval = stats.fisher_exact(table, alternative='greater')
        except Exception:
            pval = 1.0
        dom_name = COG_CODE_NAMES.get(dominant, '-')
        results.append({
            'community': c,
            'size': n,
            'dominant_cog': dominant,
            'dominant_cog_name': dom_name,
            'dominant_cog_count': dom_count,
            'dominant_cog_fraction': dom_count / n if n > 0 else 0,
            'pvalue': pval,
            'cog_distribution': json.dumps(cog_counts),
        })
    df = pd.DataFrame(results)
    # FDR-correct the COG dominance p-values
    from statsmodels.stats.multitest import multipletests
    if len(df) > 1:
        _, df['padj_cog'], _, _ = multipletests(df['pvalue'], method='fdr_bh')
    else:
        df['padj_cog'] = df['pvalue']
    n_sig = (df['padj_cog'] < 0.05).sum()
    print(f"  COG characterization: {len(df)} communities annotated, {n_sig} with significant COG enrichment (padj < 0.05)")
    return df


# ═══════════════════════════════════════════════════════════════════════
# STEP 7: Visualization (SVG figures + PyVis HTML)
# ═══════════════════════════════════════════════════════════════════════

def create_centrality_plots(centrality_df: pd.DataFrame, output_dir: str,
                            top_n: int):
    """Create centrality distribution plots + top hubs bar chart."""
    fig, axes = plt.subplots(2, 2, figsize=(14, 10))

    # Degree distribution
    axes[0, 0].hist(centrality_df['degree'], bins=30, color='#0279EE', alpha=0.7, edgecolor='white')
    axes[0, 0].set_xlabel('Degree Centrality')
    axes[0, 0].set_ylabel('Count')
    axes[0, 0].set_title('Degree Distribution')

    # Betweenness distribution
    axes[0, 1].hist(centrality_df['betweenness'], bins=30, color='#FF9400', alpha=0.7, edgecolor='white')
    axes[0, 1].set_xlabel('Betweenness Centrality')
    axes[0, 1].set_ylabel('Count')
    axes[0, 1].set_title('Betweenness Distribution')

    # Closeness distribution
    axes[1, 0].hist(centrality_df['closeness'], bins=30, color='#75A025', alpha=0.7, edgecolor='white')
    axes[1, 0].set_xlabel('Closeness Centrality')
    axes[1, 0].set_ylabel('Count')
    axes[1, 0].set_title('Closeness Distribution')

    # Top hubs bar chart
    top_hubs = centrality_df.head(top_n)
    colors = ['#FF9400' if ess else '#0279EE' for ess in top_hubs['is_essential']]
    axes[1, 1].barh(range(len(top_hubs)), top_hubs['betweenness'], color=colors, edgecolor='white')
    axes[1, 1].set_yticks(range(len(top_hubs)))
    axes[1, 1].set_yticklabels([f"{r['gene_name']}" for _, r in top_hubs.iterrows()], fontsize=8)
    axes[1, 1].set_xlabel('Betweenness Centrality')
    axes[1, 1].set_title(f'Top {top_n} Hub Genes')
    axes[1, 1].invert_yaxis()

    plt.tight_layout()
    path = os.path.join(output_dir, 'figures', 'centrality_distributions.svg')
    fig.savefig(path, format='svg', bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved: {path}")


def create_community_plot(G: nx.MultiDiGraph, Gu: nx.Graph,
                          comm_map: Dict[str, int], centrality_df: pd.DataFrame,
                          output_dir: str):
    """Create community overview network layout."""
    fig, ax = plt.subplots(figsize=(14, 10))
    pos = nx.spring_layout(Gu, k=1.5/np.sqrt(Gu.number_of_nodes()), seed=42, iterations=50)

    # Node colors by community
    n_communities = len(set(comm_map.values()))
    palette = sns.color_palette('husl', n_communities)
    node_colors = [palette[comm_map.get(n, 0) % n_communities] for n in Gu.nodes()]

    # Node sizes by betweenness
    betw_dict = dict(zip(centrality_df['locus_tag'], centrality_df['betweenness']))
    max_betw = max(betw_dict.values()) if betw_dict else 1
    node_sizes = [50 + 300 * (betw_dict.get(n, 0) / max_betw) for n in Gu.nodes()]

    # Draw edges
    nx.draw_networkx_edges(Gu, pos, ax=ax, alpha=0.15, edge_color='gray')

    # Draw nodes
    nx.draw_networkx_nodes(Gu, pos, ax=ax, node_color=node_colors,
                           node_size=node_sizes, alpha=0.8, edgecolors='white', linewidths=0.5)

    # Label top hubs
    top_hubs = centrality_df.head(15)['locus_tag'].tolist()
    labels = {n: centrality_df[centrality_df['locus_tag'] == n].iloc[0]['gene_name']
              for n in top_hubs if n in pos}
    nx.draw_networkx_labels(Gu, pos, labels, ax=ax, font_size=7, font_weight='bold')

    ax.set_title(f'Network Communities (Louvain, {n_communities} communities)', fontsize=14)
    ax.axis('off')
    plt.tight_layout()
    path = os.path.join(output_dir, 'figures', 'community_overview.svg')
    fig.savefig(path, format='svg', bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved: {path}")


def create_de_enrichment_heatmap(de_enrichment: pd.DataFrame, output_dir: str):
    """Create DE enrichment heatmap by community."""
    fig, ax = plt.subplots(figsize=(10, max(6, len(de_enrichment) * 0.4)))
    df = de_enrichment.sort_values('pvalue')
    colors = [-np.log10(max(p, 1e-10)) for p in df['padj']]
    bar_colors = ['#cc2222' if p else '#999999' for p in df['is_perturbed']]
    ax.barh(range(len(df)), colors, color=bar_colors, edgecolor='white')
    ax.set_yticks(range(len(df)))
    ax.set_yticklabels([f"Comm {int(r['community'])} (n={int(r['size'])}, DE={int(r['de_count'])})"
                        for _, r in df.iterrows()], fontsize=8)
    ax.set_xlabel('-log10(adjusted p-value)')
    ax.set_title('DE Gene Enrichment by Community')
    ax.axvline(x=-np.log10(0.05), color='black', linestyle='--', alpha=0.5, label='padj = 0.05')
    ax.legend()
    ax.invert_yaxis()
    plt.tight_layout()
    path = os.path.join(output_dir, 'figures', 'de_enrichment_heatmap.svg')
    fig.savefig(path, format='svg', bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved: {path}")


def create_edge_composition_plot(G: nx.MultiDiGraph, comm_map: Dict[str, int],
                                 output_dir: str):
    """Create edge layer composition per community."""
    # Count edges per layer per community
    comm_edges = {}
    for u, v, d in G.edges(data=True):
        cu = comm_map.get(u)
        cv = comm_map.get(v)
        layer = d.get('layer', 'unknown')
        if cu is not None:
            key = (cu, layer)
            comm_edges[key] = comm_edges.get(key, 0) + 1
        if cv is not None and cv != cu:
            key = (cv, layer)
            comm_edges[key] = comm_edges.get(key, 0) + 1

    communities = sorted(set(comm_map.values()))
    layers = ['genetic_interaction', 'suppression', 'ppi']
    layer_colors = {'genetic_interaction': '#0279EE', 'suppression': '#FF9400', 'ppi': '#75A025'}

    fig, ax = plt.subplots(figsize=(12, 6))
    x = np.arange(len(communities))
    bottom = np.zeros(len(communities))
    for layer in layers:
        counts = [comm_edges.get((c, layer), 0) for c in communities]
        ax.bar(x, counts, bottom=bottom, label=layer, color=layer_colors[layer], edgecolor='white')
        bottom += np.array(counts)
    ax.set_xticks(x)
    ax.set_xticklabels([f'C{c}' for c in communities], fontsize=8)
    ax.set_xlabel('Community')
    ax.set_ylabel('Edge Count')
    ax.set_title('Edge Layer Composition per Community')
    ax.legend()
    plt.tight_layout()
    path = os.path.join(output_dir, 'figures', 'edge_composition.svg')
    fig.savefig(path, format='svg', bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved: {path}")


def create_pyvis_html(G: nx.MultiDiGraph, comm_map: Dict[str, int],
                      centrality_df: pd.DataFrame, output_dir: str):
    """Create interactive PyVis HTML visualization."""
    from pyvis.network import Network
    net = Network(height='700px', width='100%', bgcolor='#ffffff',
                  font_color='black', directed=True)
    net.toggle_physics(True)

    betw_dict = dict(zip(centrality_df['locus_tag'], centrality_df['betweenness']))
    max_betw = max(betw_dict.values()) if betw_dict else 1
    n_comm = len(set(comm_map.values()))
    palette = sns.color_palette('husl', n_comm).as_hex()

    # Add nodes
    for node in G.nodes():
        attrs = G.nodes[node]
        size = 10 + 30 * (betw_dict.get(node, 0) / max_betw)
        color = palette[comm_map.get(node, 0) % n_comm]
        label = attrs.get('gene_name', node)
        title = (f"Gene: {label}\nLocus: {node}\n"
                 f"Betweenness: {betw_dict.get(node, 0):.4f}\n"
                 f"COG: {attrs.get('cog_category', '-')}\n"
                 f"logFC: {attrs.get('logFC', 'N/A')}\n"
                 f"FDR: {attrs.get('FDR', 'N/A')}\n"
                 f"Essential: {attrs.get('is_essential', False)}\n"
                 f"Community: {comm_map.get(node, 'N/A')}")
        net.add_node(node, label=label, size=size, color=color, title=title)

    # Add edges
    edge_colors = {'genetic_interaction': '#888888', 'suppression': '#FF9400', 'ppi': '#0279EE'}
    for u, v, d in G.edges(data=True):
        layer = d.get('layer', 'unknown')
        net.add_edge(u, v, color=edge_colors.get(layer, '#ccc'), width=1,
                     title=f"{layer}: {d.get('interaction_class', '')}")

    path = os.path.join(output_dir, 'interactive_network.html')
    net.write_html(path, notebook=False)
    print(f"  Saved: {path}")


# ═══════════════════════════════════════════════════════════════════════
# STEP 8: Community Webpage Generation
# ═══════════════════════════════════════════════════════════════════════

def _generate_venn_data(G: nx.MultiDiGraph, config: Config):
    """Generate venn_data.js with 4-study GENETIC INTERACTION overlap (UpSet plot).

    Compares signed genetic interactions (gene pairs) across studies, unified in
    SPD_ (D39W) locus-tag space via orthology mapping:
      1. Dual Tn-seq (this study) — Table S3 hits + Table S1 |z|>4, sign from z
      2. RB Tn-seq (this study) — Table S4 t-values, 71 query KOs x all genes,
         |t|>3 (median across replicate strains), sign from t
      3. Veening CRISPRi (single + dual) — pre-computed signed GI pairs from
         the Dual-CRISPRi R pipeline (data/veening_gi_pairs.json), SPV_→SPD_
      4. Opijnen Tn-seq (all studies) — CONTEXT set: a pair is "in" Opijnen when
         both genes were assayed in TIGR4 (SP_→SPD_); Tn-seq is gene x condition,
         not gene x gene, so it carries no sign.

    Item identity = (frozenset({gene1, gene2}), sign): an interaction is shared
    between GI studies only when the same pair has the same sign.
    """
    from itertools import combinations
    from collections import defaultdict

    website_dir = os.path.join(config.output_dir, 'website')
    data_dir = config.data_dir

    # ── Load orthology mappings ──
    with open(os.path.join(data_dir, 'sp_to_spd_mapping.json')) as f:
        sp_map = json.load(f)
    with open(os.path.join(data_dir, 'tigr4_genes.json')) as f:
        tigr4_genes = json.load(f)

    # Gene-name lookup: network nodes first, then TIGR4 names via mapping
    spd_names = {}
    for node in G.nodes():
        name = G.nodes[node].get('gene_name', '')
        if name:
            spd_names[node] = name
    for sp_id, spd_id in sp_map.items():
        if sp_id in tigr4_genes and tigr4_genes[sp_id]:
            spd_names.setdefault(spd_id, tigr4_genes[sp_id])

    # ── 1. Dual Tn-seq signed interactions ──
    dual_items = {}  # (frozenset, sign) -> max |z|
    wb = openpyxl.load_workbook(config.table_s3, read_only=True)
    ws = wb['confident_hit_table']
    rows = list(ws.iter_rows(min_row=12, values_only=True))
    wb.close()
    header = rows[0]
    zcol = header.index('zStrains(10-90%ORF)')
    for row in rows[1:]:
        g1, g2 = str(row[1]).strip(), str(row[2]).strip()
        if not g1 or not g2 or g1 == 'None' or g2 == 'None':
            continue
        try:
            z = float(row[zcol])
        except (TypeError, ValueError):
            continue
        sign = 'positive' if z > 0 else 'negative'
        key = (frozenset({g1.upper(), g2.upper()}), sign)
        dual_items[key] = max(dual_items.get(key, 0), abs(z))

    wb = openpyxl.load_workbook(config.table_s1, read_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i <= 16:
                continue
            g1 = str(row[1]).strip() if row[1] else ''
            g2 = str(row[2]).strip() if row[2] else ''
            if not g1 or not g2 or g1 == 'None' or g2 == 'None':
                continue
            try:
                z = float(row[13])  # zStrains(10-90%ORF)
            except (TypeError, ValueError):
                continue
            if abs(z) <= config.z_threshold:
                continue
            sign = 'positive' if z > 0 else 'negative'
            key = (frozenset({g1.upper(), g2.upper()}), sign)
            dual_items[key] = max(dual_items.get(key, 0), abs(z))
    wb.close()
    n_pos = sum(1 for k in dual_items if k[1] == 'positive')
    print(f"  Venn — Dual Tn-seq: {len(dual_items)} signed GIs "
          f"({n_pos} pos, {len(dual_items)-n_pos} neg)")

    # ── 2. RB Tn-seq signed interactions (1-to-all: query KO x target) ──
    RB_T_THRESHOLD = 3.0
    wb = openpyxl.load_workbook(config.table_s4, read_only=True)
    ws = wb['t values']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    ko_genes_row = rows[9][1:]  # row 10: KO gene per strain column
    col_ko = {}
    for j, ko in enumerate(ko_genes_row, start=1):
        if ko and ko != 'WT':
            col_ko[j] = str(ko).strip().upper()
    rb_tvals = defaultdict(list)
    for row in rows[12:]:
        target = row[0]
        if target is None:
            continue
        target = str(target).strip().upper()
        for j, ko in col_ko.items():
            if j < len(row):
                v = row[j]
                if v is not None and v != 'N/A' and v != '':
                    try:
                        rb_tvals[(ko, target)].append(float(v))
                    except (TypeError, ValueError):
                        pass
    rb_items = {}
    for (ko, target), vals in rb_tvals.items():
        t = float(np.median(vals))
        if abs(t) > RB_T_THRESHOLD and ko != target:
            sign = 'positive' if t > 0 else 'negative'
            key = (frozenset({ko, target}), sign)
            rb_items[key] = max(rb_items.get(key, 0), abs(t))
    n_pos = sum(1 for k in rb_items if k[1] == 'positive')
    print(f"  Venn — RB Tn-seq: {len(rb_items)} signed GIs "
          f"({n_pos} pos, {len(rb_items)-n_pos} neg) from {len(set(col_ko.values()))} query KOs")

    # ── 3. Veening CRISPRi signed interactions (pre-computed, SPV_→SPD_) ──
    with open(os.path.join(data_dir, 'veening_gi_pairs.json')) as f:
        veening_raw = json.load(f)
    veening_items = {}
    for rec in veening_raw:
        pair = frozenset(rec['pair'])
        key = (pair, rec['sign'])
        veening_items[key] = max(veening_items.get(key, 0), rec['score'])
    n_pos = sum(1 for k in veening_items if k[1] == 'positive')
    print(f"  Venn — Veening CRISPRi: {len(veening_items)} signed GIs "
          f"({n_pos} pos, {len(veening_items)-n_pos} neg)")

    # ── 4. Opijnen Tn-seq context set (both genes assayed in TIGR4) ──
    opijnen_genes = set(sp_map[sp] for sp in tigr4_genes if sp in sp_map)
    all_items = set(dual_items) | set(rb_items) | set(veening_items)
    opijnen_items = {k for k in all_items if k[0] <= opijnen_genes}
    print(f"  Venn — Opijnen Tn-seq (context): {len(opijnen_items)} items "
          f"with both genes assayed (of {len(opijnen_genes)} testable genes)")

    # ── 5. Compute all 15 intersections ──
    study_sets = {
        'Dual Tn-seq': set(dual_items),
        'RB Tn-seq': set(rb_items),
        'Veening CRISPRi': set(veening_items),
        'Opijnen Tn-seq': opijnen_items,
    }
    method_ids = list(study_sets.keys())

    def item_scores(key):
        scores = {}
        if key in dual_items:
            scores['Dual Tn-seq'] = round(dual_items[key], 2)
        if key in rb_items:
            scores['RB Tn-seq'] = round(rb_items[key], 2)
        if key in veening_items:
            scores['Veening CRISPRi'] = round(veening_items[key], 2)
        return scores

    MAX_ITEMS = 500  # cap per intersection for browser-friendly file size
    intersections = []
    for r in range(1, 5):
        for combo in combinations(range(4), r):
            sets_involved = [method_ids[i] for i in combo]
            inter = study_sets[sets_involved[0]]
            for s in sets_involved[1:]:
                inter = inter & study_sets[s]
            others = [study_sets[method_ids[i]] for i in range(4) if i not in combo]
            excl = inter
            for o in others:
                excl = excl - o
            if len(inter) == 0:
                continue
            # Build item details, sorted by max |score| across studies
            items = []
            for key in excl:
                pair, sign = key
                g = sorted(pair)
                scores = item_scores(key)
                items.append({
                    'n1': spd_names.get(g[0], g[0]),
                    'n2': spd_names.get(g[1], g[1]),
                    'l1': g[0],
                    'l2': g[1],
                    'sign': sign,
                    'scores': scores,
                    'methods': [m for m in method_ids if key in study_sets[m]],
                    '_maxscore': max(scores.values()) if scores else 0,
                })
            items.sort(key=lambda x: -x['_maxscore'])
            truncated = len(items) > MAX_ITEMS
            items = items[:MAX_ITEMS]
            for it in items:
                del it['_maxscore']
            intersections.append({
                'sets': sets_involved,
                'set_ids': list(combo),
                'size': len(inter),
                'exclusive': len(excl),
                'truncated': truncated,
                'items': items,
            })

    method_genes = {m: len(study_sets[m]) for m in method_ids}
    venn_obj = {
        'methods': method_ids,
        'method_totals': method_genes,
        'method_genes': method_genes,
        'set_types': {
            'Dual Tn-seq': 'interaction',
            'RB Tn-seq': 'interaction',
            'Veening CRISPRi': 'interaction',
            'Opijnen Tn-seq': 'context',
        },
        'score_labels': {
            'Dual Tn-seq': '|z|',
            'RB Tn-seq': '|t|',
            'Veening CRISPRi': '|ε|',
        },
        'intersections': intersections,
    }

    venn_js = f"""// PneumoGI Venn/UpSet Data — Generated by grn_pipeline.py
// 4-study GENETIC INTERACTION overlap (signed gene pairs, SPD_ orthology space)
// Opijnen Tn-seq = context set (both genes assayed); other sets = signed GIs
// Generated {time.strftime('%Y-%m-%d')}
const VENN_DATA = {json.dumps(venn_obj, indent=None)};
"""
    venn_path = os.path.join(website_dir, 'venn_data.js')
    with open(venn_path, 'w') as f:
        f.write(venn_js)
    total_items = sum(len(ix['items']) for ix in intersections)
    print(f"  Saved: {venn_path} ({len(intersections)} intersections, "
          f"{total_items} interaction items)")


def generate_webpage(G: nx.MultiDiGraph, centrality_df: pd.DataFrame,
                     comm_map: Dict[str, int], de_enrichment: pd.DataFrame,
                     cog_chars: pd.DataFrame, config: Config):
    """Generate standalone community webpage (index.html + data.js)."""
    website_dir = os.path.join(config.output_dir, 'website')
    os.makedirs(website_dir, exist_ok=True)

    # ── Generate data.js ──
    gi_entries = []
    edge_id = 0
    for u, v, d in G.edges(data=True):
        edge_id += 1
        ua = G.nodes[u]
        va = G.nodes[v]
        gi_entries.append({
            'id': f'GI_{edge_id:05d}',
            'gene_a': ua.get('gene_name', u),
            'gene_a_locus': u,
            'gene_b': va.get('gene_name', v),
            'gene_b_locus': v,
            'class': d.get('interaction_class', 'unknown'),
            'subtype': d.get('subtype', ''),
            'method': d.get('source', ''),
            'score': str(d.get('zscore', d.get('score', ''))),
            'strain': 'D39W',
            'condition': 'in vitro BHI',
            'func_a': ua.get('product', '-'),
            'func_b': va.get('product', '-'),
            'cog_a': ua.get('cog_category', '-'),
            'cog_b': va.get('cog_category', '-'),
            'validation': d.get('validation', 'Untested'),
            'doi': '10.1126/science.adt7685' if 'Dual Tn-seq' in d.get('source', '') else '',
            'community_a': comm_map.get(u, -1),
            'community_b': comm_map.get(v, -1),
        })

    node_attrs = {}
    for node in G.nodes():
        a = G.nodes[node]
        betw = dict(zip(centrality_df['locus_tag'], centrality_df['betweenness'])).get(node, 0)
        deg = dict(zip(centrality_df['locus_tag'], centrality_df['degree'])).get(node, 0)
        node_attrs[node] = {
            'gene_name': a.get('gene_name', node),
            'cog_category': a.get('cog_category', '-'),
            'logFC': a.get('logFC'),
            'FDR': a.get('FDR'),
            'is_de': a.get('is_de', False),
            'is_essential': a.get('is_essential', False),
            'fitness_median': a.get('fitness_median'),
            'community': comm_map.get(node, -1),
            'betweenness': round(betw, 6),
            'degree': round(deg, 4),
            'product': a.get('product', '-'),
        }

    # Canonical map
    canonical = {}
    for locus, name in node_attrs.items():
        if name['gene_name'] and name['gene_name'] != locus:
            canonical[name['gene_name'].lower()] = locus
            canonical[name['gene_name']] = locus
        canonical[locus.lower()] = locus
        canonical[locus] = locus

    # Stats
    n_neg = sum(1 for e in gi_entries if e['class'] == 'negative')
    n_pos = sum(1 for e in gi_entries if e['class'] == 'positive')
    n_sup = sum(1 for e in gi_entries if e['class'] == 'suppression')
    n_ppi = sum(1 for e in gi_entries if e['class'] == 'PPI')
    n_comm = len(set(comm_map.values()))
    n_hubs = min(config.top_hubs, len(centrality_df))

    # Community annotations for webpage
    comm_merged = de_enrichment.merge(cog_chars, on='community', how='left')
    comm_annotations = []
    for _, row in comm_merged.iterrows():
        comm_annotations.append({
            'community': int(row['community']),
            'size': int(row.get('size_x', row.get('size', 0))),
            'de_count': int(row['de_count']),
            'de_expected': round(float(row['de_expected']), 1),
            'de_enrichment': round(float(row['de_enrichment']), 3),
            'de_padj': round(float(row['padj']), 4),
            'is_perturbed': bool(row['is_perturbed']),
            'dominant_cog': str(row.get('dominant_cog', '-')),
            'dominant_cog_name': str(row.get('dominant_cog_name', '-')),
            'dominant_cog_fraction': round(float(row.get('dominant_cog_fraction', 0)), 3),
            'cog_padj': round(float(row.get('padj_cog', 1.0)), 4),
        })
    comm_annotations.sort(key=lambda x: x['community'])

    data_js = f"""// PneumoGI Network Database — Generated by grn_pipeline.py
// Generated {time.strftime('%Y-%m-%d')}
const GI_DATABASE = {json.dumps(gi_entries, indent=None)};
const NODE_ATTRIBUTES = {json.dumps(node_attrs, indent=None)};
const CANONICAL_MAP = {json.dumps(canonical, indent=None)};
const COMMUNITY_ANNOTATIONS = {json.dumps(comm_annotations, indent=None)};
const STATS = {{
  total: {len(gi_entries)},
  negative: {n_neg},
  positive: {n_pos},
  suppression: {n_sup},
  ppi: {n_ppi},
  communities: {n_comm},
  hubs: {n_hubs},
  nodes: {G.number_of_nodes()},
}};
"""

    data_path = os.path.join(website_dir, 'data.js')
    with open(data_path, 'w') as f:
        f.write(data_js)
    print(f"  Saved: {data_path} ({len(gi_entries)} interactions)")

    # ── Generate index.html ──
    html = _generate_html(n_comm, n_hubs)
    html_path = os.path.join(website_dir, 'index.html')
    with open(html_path, 'w') as f:
        f.write(html)
    print(f"  Saved: {html_path}")

    # ── Generate venn_data.js ──
    _generate_venn_data(G, config)


def _generate_html(n_comm: int, n_hubs: int) -> str:
    """Generate the HTML page with inline CSS and JS."""
    return r'''<!DOCTYPE html>
<html lang="en" data-theme="dark">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0"/>
<title>PneumoGI — Genetic Interaction Network Database</title>
<link rel="preconnect" href="https://fonts.googleapis.com"/>
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin/>
<link href="https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;600&family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet"/>
<script src="data.js"></script>
<style>
:root{--text-xs:clamp(.75rem,.7rem+.25vw,.875rem);--text-sm:clamp(.875rem,.8rem+.35vw,1rem);--text-base:clamp(1rem,.95rem+.25vw,1.125rem);--text-lg:clamp(1.125rem,1rem+.75vw,1.5rem);--text-xl:clamp(1.5rem,1.2rem+1.25vw,2.25rem);--text-2xl:clamp(2rem,1.2rem+2.5vw,3.5rem);--font-d:'JetBrains Mono',monospace;--font-b:'Inter',system-ui,sans-serif;--r-s:4px;--r-m:8px;--r-l:12px;--t:.18s ease}
:root,[data-theme="light"]{--bg:#f4f4f0;--srf:#fff;--elv:#f0efea;--bd:rgba(0,0,0,.1);--bds:rgba(0,0,0,.2);--tx:#1a1a18;--tx2:#555550;--txm:#888880;--ac:#0066cc;--acs:rgba(0,102,204,.1);--neg:#cc2222;--pos:#0f7a3c;--sup:#cc6600;--ppi:#6666cc;--sh:0 1px 3px rgba(0,0,0,.08)}
[data-theme="dark"]{--bg:#0d0e0f;--srf:#161718;--elv:#1e2022;--bd:rgba(255,255,255,.08);--bds:rgba(255,255,255,.15);--tx:#e8e9ea;--tx2:#9a9b9d;--txm:#5e6062;--ac:#4d9fff;--acs:rgba(77,159,255,.12);--neg:#ff5a5a;--pos:#4dcc80;--sup:#ffaa44;--ppi:#9999ff;--sh:0 1px 3px rgba(0,0,0,.4)}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
body{font-family:var(--font-b);font-size:var(--text-base);background:var(--bg);color:var(--tx);line-height:1.6;-webkit-font-smoothing:antialiased;transition:background var(--t),color var(--t)}
.hdr{position:sticky;top:0;z-index:100;background:var(--srf);border-bottom:1px solid var(--bd);backdrop-filter:blur(12px)}
.hdr-i{max-width:1400px;margin:0 auto;padding:1rem 1.5rem;display:flex;align-items:center;gap:1rem}
.logo{display:flex;align-items:center;gap:.75rem;text-decoration:none;color:var(--tx)}
.logo-t{font-family:var(--font-d);font-size:var(--text-lg);font-weight:600;letter-spacing:-.03em}
.logo-t span{color:var(--ac)}
.logo-s{font-size:var(--text-xs);color:var(--txm);text-transform:uppercase;letter-spacing:.04em}
.nav{display:flex;gap:.5rem;margin-left:auto;align-items:center}
.nb{background:none;border:1px solid var(--bd);border-radius:var(--r-m);padding:.5rem 1rem;font-family:var(--font-b);font-size:var(--text-sm);color:var(--tx2);cursor:pointer;transition:all var(--t)}
.nb:hover{background:var(--elv);color:var(--tx)}
.nb.act{background:var(--acs);border-color:var(--ac);color:var(--ac)}
.tb{background:none;border:1px solid var(--bd);border-radius:var(--r-m);width:36px;height:36px;cursor:pointer;display:flex;align-items:center;justify-content:center;color:var(--tx2);transition:all var(--t)}
.tb:hover{background:var(--elv);color:var(--tx)}
.hero{background:var(--srf);border-bottom:1px solid var(--bd);padding:2.5rem 1.5rem 2rem}
.hero-i{max-width:1400px;margin:0 auto}
.hero-t{font-family:var(--font-d);font-size:var(--text-2xl);font-weight:600;letter-spacing:-.04em;line-height:1.1;margin-bottom:.75rem}
.hero-t em{color:var(--ac);font-style:normal}
.hero-s{color:var(--tx2);max-width:640px;font-size:var(--text-sm);line-height:1.7;margin-bottom:2rem}
.sr{display:flex;flex-wrap:wrap;gap:1rem}
.sc{background:var(--elv);border:1px solid var(--bd);border-radius:var(--r-l);padding:1rem 1.5rem;min-width:120px}
.sn{font-family:var(--font-d);font-size:var(--text-xl);font-weight:600;color:var(--ac);letter-spacing:-.03em;line-height:1}
.sl{font-size:var(--text-xs);color:var(--txm);text-transform:uppercase;letter-spacing:.05em;margin-top:.25rem}
.tab{display:none;max-width:1400px;margin:0 auto;padding:1.5rem}
.tab.act{display:block}
.fbar{display:flex;flex-wrap:wrap;gap:.5rem;margin-bottom:1rem;align-items:center}
.sel,.inp{background:var(--srf);border:1px solid var(--bd);border-radius:var(--r-m);padding:.5rem .75rem;font-family:var(--font-b);font-size:var(--text-sm);color:var(--tx)}
.sel:focus,.inp:focus{outline:none;border-color:var(--ac)}
table{width:100%;border-collapse:collapse;font-size:var(--text-xs)}
th,td{padding:.5rem .75rem;text-align:left;border-bottom:1px solid var(--bd)}
th{cursor:pointer;white-space:nowrap;color:var(--tx2);font-weight:600;text-transform:uppercase;letter-spacing:.03em;font-size:.7rem}
th:hover{color:var(--ac)}
tr:hover{background:var(--elv)}
.badge{display:inline-block;padding:.15rem .5rem;border-radius:var(--r-s);font-size:.65rem;font-weight:600;text-transform:uppercase}
.b-neg{background:rgba(204,34,34,.15);color:var(--neg)}
.b-pos{background:rgba(15,122,60,.15);color:var(--pos)}
.b-sup{background:rgba(204,102,0,.15);color:var(--sup)}
.b-ppi{background:rgba(102,102,204,.15);color:var(--ppi)}
.b-ess{background:rgba(255,200,0,.15);color:#ffc800}
.b-de{background:rgba(253,155,237,.15);color:#FD9BED}
.pager{display:flex;gap:.5rem;justify-content:center;margin-top:1rem;align-items:center}
.pgbtn{background:var(--elv);border:1px solid var(--bd);border-radius:var(--r-s);padding:.4rem .8rem;cursor:pointer;font-size:var(--text-sm);color:var(--tx)}
.pgbtn:hover{background:var(--ac);color:#fff}
.nw-wrap{display:flex;gap:1rem;flex-wrap:wrap}
.nw-canvas{flex:1;min-width:0;background:var(--srf);border:1px solid var(--bd);border-radius:var(--r-l);overflow:hidden}
canvas{display:block}
.nw-side{width:240px;display:flex;flex-direction:column;gap:1rem;transition:width var(--t),opacity var(--t)}
.nw-side.collapsed{display:none}
.nw-legend{background:var(--srf);border:1px solid var(--bd);border-radius:var(--r-l);padding:1rem}
.lg-item{display:flex;align-items:center;gap:.5rem;font-size:var(--text-xs);margin-bottom:.4rem}
.lg-dot{width:12px;height:12px;border-radius:50%;flex-shrink:0}
.lg-line{width:20px;height:3px;flex-shrink:0;border-radius:2px}
.cbar{height:4px;border-radius:2px;background:var(--ac);margin-top:2px}
.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(300px,1fr));gap:1rem;margin-bottom:1.5rem}
.card{background:var(--srf);border:1px solid var(--bd);border-radius:var(--r-l);padding:1.5rem}
.card h3{font-family:var(--font-d);font-size:var(--text-sm);font-weight:600;margin-bottom:1rem;color:var(--tx2)}
.btn-dl{display:inline-flex;align-items:center;gap:.5rem;background:var(--ac);color:#fff;border:none;border-radius:var(--r-m);padding:.6rem 1.2rem;font-family:var(--font-b);font-size:var(--text-sm);cursor:pointer;transition:all var(--t)}
.btn-dl:hover{opacity:.85}
.ftr{border-top:1px solid var(--bd);padding:1.5rem;text-align:center;color:var(--txm);font-size:var(--text-xs)}
.sig{color:var(--pos);font-weight:600}
.nosig{color:var(--txm)}
</style>
</head>
<body>
<div class="hdr"><div class="hdr-i">
<a class="logo" href="#"><div class="logo-t">Pneumo<span>GI</span></div><div class="logo-s">Genetic Interaction Network</div></a>
<div class="nav">
<button class="nb act" onclick="showTab('table',this)">Interaction Table</button>
<button class="nb" onclick="showTab('network',this)">Network View</button>
<button class="nb" onclick="showTab('summary',this)">Summary &amp; Hubs</button>
<a class="nb" href="venn.html" style="text-decoration:none">Venn Diagram</a>
<button class="tb" onclick="toggleTheme()" id="themeBtn">◐</button>
</div>
</div></div>

<div class="hero"><div class="hero-i">
<h1 class="hero-t"><em>Streptococcus pneumoniae</em> Genetic Interaction Network</h1>
<p class="hero-s">A multi-layer network integrating Dual Tn-seq genetic interactions, Sup-seq suppression edges, STRING PPI, and RNA-seq differential expression. Built from data by Zik et al. 2025 and Ng et al. 2025, Sham Lab NUS.</p>
<div class="sr" id="statsRow"></div>
</div></div>

<!-- TAB 1: TABLE -->
<div class="tab act" id="tab-table">
<div class="fbar">
<input class="inp" type="text" id="search" placeholder="Search gene name or locus..." oninput="renderTable()" style="flex:1;min-width:200px"/>
<select class="sel" id="fClass" onchange="renderTable()"><option>All classes</option><option>negative</option><option>positive</option><option>suppression</option><option>PPI</option></select>
<select class="sel" id="fMethod" onchange="renderTable()"><option>All methods</option><option>Dual Tn-seq</option><option>Dual Tn-seq (strict)</option><option>Sup-seq</option><option>STRING</option></select>
<select class="sel" id="fVal" onchange="renderTable()"><option>All validation</option><option>Untested</option><option>Published</option><option>Computational</option></select>
</div>
<div style="overflow-x:auto"><table id="giTable"><thead><tr>
<th onclick="sortTable('id')">ID</th><th onclick="sortTable('gene_a')">Gene A</th><th onclick="sortTable('gene_b')">Gene B</th>
<th onclick="sortTable('class')">Class</th><th onclick="sortTable('subtype')">Subtype</th><th onclick="sortTable('method')">Method</th>
<th onclick="sortTable('score')">Score</th><th onclick="sortTable('cog_a')">COG A</th><th onclick="sortTable('cog_b')">COG B</th>
<th onclick="sortTable('validation')">Validation</th></tr></thead><tbody id="tbody"></tbody></table></div>
<div class="pager" id="pager"></div>
</div>

<!-- TAB 2: NETWORK -->
<div class="tab" id="tab-network">
<div class="fbar">
<input class="inp" type="text" id="nwSearch" placeholder="Search gene..." oninput="filterNetwork()" style="flex:1;min-width:200px"/>
<select class="sel" id="nwMethod" onchange="filterNetwork()"><option>All methods</option><option>Dual Tn-seq</option><option>Dual Tn-seq (strict)</option><option>Sup-seq</option><option>STRING</option></select>
<select class="sel" id="nwClass" onchange="filterNetwork()"><option>All classes</option><option>negative</option><option>positive</option><option>suppression</option><option>PPI</option></select>
<button class="nb" onclick="clearFocus()">✕ Clear focus</button>
<button class="nb" onclick="resetLayout()">↺ Reset layout</button>
<button class="nb" onclick="toggleSidebar()" id="sbToggle">▸ Hide Panel</button>
</div>
<div class="nw-wrap">
<div class="nw-canvas"><canvas id="nwCanvas" height="600"></canvas></div>
<div class="nw-side" id="nwSide">
<div class="nw-legend">
<div style="font-family:var(--font-d);font-size:var(--text-sm);font-weight:600;margin-bottom:.5rem">Edges</div>
<div class="lg-item"><div class="lg-line" style="background:var(--neg)"></div>Negative GI</div>
<div class="lg-item"><div class="lg-line" style="background:var(--pos)"></div>Positive GI</div>
<div class="lg-item"><div class="lg-line" style="background:var(--sup)"></div>Suppression</div>
<div class="lg-item"><div class="lg-line" style="background:var(--ppi)"></div>PPI</div>
</div>
<div class="nw-legend">
<div style="font-family:var(--font-d);font-size:var(--text-sm);font-weight:600;margin-bottom:.5rem">Nodes</div>
<div style="font-size:var(--text-xs);color:var(--tx2)">Size ∝ betweenness centrality<br/>Color ∝ community</div>
</div>
<div id="nwInfo" class="nw-legend" style="display:none">
<div style="font-family:var(--font-d);font-size:var(--text-sm);font-weight:600;margin-bottom:.5rem">Focused Node</div>
<div id="nwInfoContent" style="font-size:var(--text-xs);color:var(--tx2)"></div>
</div>
</div>
</div>
<div style="text-align:center;color:var(--txm);font-size:var(--text-xs);margin-top:.5rem">Click node to focus · drag · scroll to zoom</div>
</div>

<!-- TAB 3: SUMMARY -->
<div class="tab" id="tab-summary">
<div class="grid">
<div class="card"><h3>Interactions by Method</h3><canvas id="chartMethod" width="280" height="200"></canvas></div>
<div class="card"><h3>Interaction Classes</h3><canvas id="chartClass" width="280" height="200"></canvas></div>
</div>
<div class="card" style="margin-bottom:1.5rem"><h3>Top ''' + str(n_hubs) + r''' Hub Genes (by Betweenness Centrality)</h3>
<div style="overflow-x:auto"><table id="hubTable"><thead><tr><th>Rank</th><th>Gene</th><th>Locus</th><th>Product</th><th>Betweenness</th><th>Degree</th><th>COG</th><th>Community</th><th>logFC</th><th>FDR</th><th>Essential</th></tr></thead><tbody id="hubBody"></tbody></table></div></div>
<div class="card" style="margin-bottom:1.5rem"><h3>Community Functional Characterization</h3>
<div style="font-size:var(--text-xs);color:var(--txm);margin-bottom:.75rem">DE enrichment: <span id="deSigCount">0</span>/''' + str(n_comm) + r''' communities significant (padj &lt; 0.05) · COG enrichment: <span id="cogSigCount">0</span>/''' + str(n_comm) + r''' communities significant</div>
<div style="overflow-x:auto"><table id="commTable"><thead><tr><th>Community</th><th>Size</th><th>DE Count</th><th>DE Enrich</th><th>DE padj</th><th>Dominant COG</th><th>COG Name</th><th>COG padj</th></tr></thead><tbody id="commBody"></tbody></table></div></div>
<div class="card" style="margin-bottom:1.5rem">
<h3>Data Sources</h3>
<table><thead><tr><th>Dataset</th><th>Method</th><th>Scale</th><th>Reference</th></tr></thead><tbody>
<tr><td>Dual Tn-seq</td><td>Transposon sequencing</td><td>~1M double mutants</td><td>Zik et al. 2025 (doi:10.1126/science.adt7685)</td></tr>
<tr><td>Sup-seq</td><td>Barcoded Tn-seq suppression</td><td>7 essential genes</td><td>Ng et al. 2025</td></tr>
<tr><td>STRING</td><td>PPI database</td><td>Score > 0.7</td><td>STRING v12.0, taxon 1313</td></tr>
<tr><td>RNA-seq</td><td>Transcriptomics</td><td>1987 genes</td><td>Ng et al. 2025</td></tr>
</tbody></table>
</div>
<button class="btn-dl" onclick="downloadCSV()">⬇ Download Full Database CSV</button>
</div>

<div class="ftr" id="footer"></div>

<script>
// ── STATE ──
let sortKey='id',sortAsc=true,page=0,pageSize=50;
let nwNodes=[],nwEdges=[],nwPos={},focusNode=null,dragging=false,lastMouse=null,scale=1,offsetX=0,offsetY=0;

// ── INIT ──
function init(){
  const s=STATS;
  const statsHtml=[
    ['Interactions',s.total],['Negative GIs',s.negative],['Positive GIs',s.positive],
    ['Suppression',s.suppression],['PPI edges',s.ppi],['Communities',s.communities],['Hub genes',s.hubs]
  ].map(([l,v])=>`<div class="sc"><div class="sn">${v}</div><div class="sl">${l}</div></div>`).join('');
  document.getElementById('statsRow').innerHTML=statsHtml;
  document.getElementById('footer').textContent=
    `${s.total} interactions · ${s.nodes} genes · Zik 2025 + Ng 2025 · Sham Lab NUS · Last updated ${new Date().toISOString().slice(0,7)}`;
  renderTable();renderHubs();renderCommunities();drawCharts();initNetwork();
  // URL search parameter handling (for Venn diagram gene links)
  const params=new URLSearchParams(window.location.search);
  const q=params.get('search');
  if(q){
    document.getElementById('search').value=q;
    const tableBtn=document.querySelector('.nb');
    showTab('table',tableBtn);
    renderTable();
  }
}

// ── TABS ──
function showTab(t,btn){
  document.querySelectorAll('.tab').forEach(e=>e.classList.remove('act'));
  document.querySelectorAll('.nb').forEach(e=>e.classList.remove('act'));
  document.getElementById('tab-'+t).classList.add('act');
  if(btn)btn.classList.add('act');
  if(t==='network')setTimeout(()=>{resizeCanvas();drawNetwork();},100);
}

// ── THEME ──
function toggleTheme(){
  const h=document.documentElement;
  h.setAttribute('data-theme',h.getAttribute('data-theme')==='dark'?'light':'dark');
}

// ── TABLE ──
function getFiltered(){
  const q=document.getElementById('search').value.toLowerCase();
  const fc=document.getElementById('fClass').value;
  const fm=document.getElementById('fMethod').value;
  const fv=document.getElementById('fVal').value;
  return GI_DATABASE.filter(e=>{
    if(q&&!(`${e.gene_a} ${e.gene_b} ${e.gene_a_locus} ${e.gene_b_locus}`.toLowerCase().includes(q)))return false;
    if(fc!=='All classes'&&e.class!==fc)return false;
    if(fm!=='All methods'&&!e.method.includes(fm))return false;
    if(fv!=='All validation'&&e.validation!==fv)return false;
    return true;
  });
}
function sortTable(k){if(sortKey===k)sortAsc=!sortAsc;else{sortKey=k;sortAsc=true;}renderTable();}
function renderTable(){
  let data=getFiltered();
  data.sort((a,b)=>{
    let va=a[sortKey]||'',vb=b[sortKey]||'';
    if(typeof va==='number')return sortAsc?va-vb:vb-va;
    return sortAsc?String(va).localeCompare(String(vb)):String(vb).localeCompare(String(va));
  });
  const tp=Math.ceil(data.length/pageSize);
  if(page>=tp)page=0;
  const slice=data.slice(page*pageSize,(page+1)*pageSize);
  const bc={negative:'b-neg',positive:'b-pos',suppression:'b-sup',PPI:'b-ppi'};
  document.getElementById('tbody').innerHTML=slice.map(e=>
    `<tr><td>${e.id}</td><td>${e.gene_a}</td><td>${e.gene_b}</td>
    <td><span class="badge ${bc[e.class]||''}">${e.class}</span></td>
    <td>${e.subtype}</td><td>${e.method}</td><td>${e.score}</td>
    <td>${e.cog_a}</td><td>${e.cog_b}</td><td>${e.validation}</td></tr>`
  ).join('');
  let ph=`<button class="pgbtn" onclick="page=0;renderTable()">«</button>`;
  ph+=`<button class="pgbtn" onclick="if(page>0)page--;renderTable()">‹</button>`;
  ph+=`<span style="font-size:var(--text-sm);color:var(--txm)">${page+1}/${tp||1} (${data.length})</span>`;
  ph+=`<button class="pgbtn" onclick="if(page<tp-1)page++;renderTable()">›</button>`;
  ph+=`<button class="pgbtn" onclick="page=${tp-1};renderTable()">»</button>`;
  document.getElementById('pager').innerHTML=ph;
}

// ── HUBS ──
function renderHubs(){
  const nodes=Object.entries(NODE_ATTRIBUTES).sort((a,b)=>b[1].betweenness-a[1].betweenness).slice(0,STATS.hubs);
  const maxBetw=nodes.length?nodes[0][1].betweenness:1;
  document.getElementById('hubBody').innerHTML=nodes.map(([locus,a],i)=>{
    const betwPct=Math.round((a.betweenness/maxBetw)*100);
    const deColor=a.is_de?(a.logFC<0?'var(--neg)':'var(--pos)'):'var(--txm)';
    const prod=(a.product||'-').replace(/%2C/g,',').replace(/%2c/g,',');
    return `<tr><td>${i+1}</td>
    <td><a href="javascript:void(0)" onclick="focusGene('${locus}')" style="color:var(--ac);text-decoration:none"><b>${a.gene_name}</b></a></td>
    <td>${locus}</td><td style="max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${prod}</td>
    <td>${a.betweenness.toFixed(4)}<div class="cbar" style="width:${betwPct}%"></div></td>
    <td>${a.degree.toFixed(3)}</td><td>${a.cog_category}</td><td>${a.community}</td>
    <td style="color:${deColor}">${a.logFC!==null?a.logFC.toFixed(2):'-'}</td>
    <td>${a.FDR!==null?a.FDR.toExponential(1):'-'}</td>
    <td>${a.is_essential?'<span class="badge b-ess">Essential</span>':''}</td></tr>`;
  }).join('');
}
function focusGene(locus){
  const a=NODE_ATTRIBUTES[locus]||{};
  const q=a.gene_name||locus;
  document.getElementById('search').value=q;
  const tableBtn=document.querySelectorAll('.nb')[0];
  showTab('table',tableBtn);
  renderTable();
}

// ── COMMUNITIES ──
function renderCommunities(){
  if(typeof COMMUNITY_ANNOTATIONS==='undefined')return;
  const rows=COMMUNITY_ANNOTATIONS;
  const deSig=rows.filter(r=>r.de_padj<0.05).length;
  const cogSig=rows.filter(r=>r.cog_padj<0.05).length;
  document.getElementById('deSigCount').textContent=deSig;
  document.getElementById('cogSigCount').textContent=cogSig;
  document.getElementById('commBody').innerHTML=rows.map(r=>{
    const deCls=r.de_padj<0.05?'sig':'nosig';
    const cogCls=r.cog_padj<0.05?'sig':'nosig';
    return `<tr><td>Community ${r.community}</td><td>${r.size}</td>
    <td>${r.de_count}</td><td>${r.de_enrichment.toFixed(2)}x</td>
    <td class="${deCls}">${r.de_padj<0.001?r.de_padj.toExponential(1):r.de_padj.toFixed(4)}</td>
    <td>${r.dominant_cog}</td><td>${r.dominant_cog_name}</td>
    <td class="${cogCls}">${r.cog_padj<0.001?r.cog_padj.toExponential(1):r.cog_padj.toFixed(4)}</td></tr>`;
  }).join('');
}

// ── CHARTS ──
function drawCharts(){
  const methods={};
  GI_DATABASE.forEach(e=>{methods[e.method]=(methods[e.method]||0)+1;});
  drawBarChart('chartMethod',methods);
  const classes={};
  GI_DATABASE.forEach(e=>{classes[e.class]=(classes[e.class]||0)+1;});
  drawBarChart('chartClass',classes);
}
function drawBarChart(id,data){
  const c=document.getElementById(id),ctx=c.getContext('2d');
  ctx.clearRect(0,0,c.width,c.height);
  const entries=Object.entries(data).sort((a,b)=>b[1]-a[1]);
  const max=Math.max(...entries.map(e=>e[1]));
  const colors=['#0279EE','#FF9400','#75A025','#FD9BED','#cc2222','#0f7a3c'];
  const style=getComputedStyle(document.documentElement);
  const tx2Color=style.getPropertyValue('--tx2').trim()||'#555';
  const bw=c.width/(entries.length*1.5);
  entries.forEach(([k,v],i)=>{
    const x=i*(bw*1.5)+bw*0.25;
    const h=(v/max)*(c.height-40);
    ctx.fillStyle=colors[i%colors.length];
    ctx.fillRect(x,c.height-h-20,bw,h);
    ctx.fillStyle=tx2Color;
    ctx.font='10px Inter';
    ctx.textAlign='center';
    const label=k.length>12?k.slice(0,10)+'..':k;
    ctx.fillText(label,x+bw/2,c.height-5);
    ctx.fillText(v,x+bw/2,c.height-h-25);
  });
}

// ── NETWORK ──
function initNetwork(){
  const canvas=document.getElementById('nwCanvas');
  const initW=820;
  const nodeSet=new Set();
  GI_DATABASE.forEach(e=>{nodeSet.add(e.gene_a_locus);nodeSet.add(e.gene_b_locus);});
  nwNodes=Array.from(nodeSet).map(locus=>({
    id:locus,
    label:NODE_ATTRIBUTES[locus]?.gene_name||locus,
    betweenness:NODE_ATTRIBUTES[locus]?.betweenness||0,
    community:NODE_ATTRIBUTES[locus]?.community||0,
    isEssential:NODE_ATTRIBUTES[locus]?.is_essential||false,
  }));
  nwEdges=GI_DATABASE.map(e=>({
    source:e.gene_a_locus,target:e.gene_b_locus,
    class:e.class,method:e.method,
  }));
  nwNodes.forEach((n,i)=>{
    nwPos[n.id]={x:initW/2+Math.cos(i*0.5)*200,y:300+Math.sin(i*0.5)*200,vx:0,vy:0};
  });
  canvas.addEventListener('mousedown',e=>{const r=canvas.getBoundingClientRect();dragging=true;lastMouse={x:e.clientX-r.left,y:e.clientY-r.top};});
  canvas.addEventListener('mousemove',e=>{
    const r=canvas.getBoundingClientRect();
    if(dragging&&lastMouse){offsetX+=e.clientX-r.left-lastMouse.x;offsetY+=e.clientY-r.top-lastMouse.y;lastMouse={x:e.clientX-r.left,y:e.clientY-r.top};drawNetwork();}
  });
  canvas.addEventListener('mouseup',()=>{dragging=false;lastMouse=null;});
  canvas.addEventListener('mouseleave',()=>{dragging=false;});
  canvas.addEventListener('wheel',e=>{e.preventDefault();const f=e.deltaY>0?0.9:1.1;scale*=f;drawNetwork();});
  canvas.addEventListener('click',e=>{
    const r=canvas.getBoundingClientRect();
    const mx=(e.clientX-r.left-offsetX)/scale,my=(e.clientY-r.top-offsetY)/scale;
    let clicked=null;
    for(const n of nwNodes){
      const p=nwPos[n.id];const dx=mx-p.x,dy=my-p.y;
      const rad=5+15*(n.betweenness/Math.max(...nwNodes.map(x=>x.betweenness||0)));
      if(dx*dx+dy*dy<rad*rad){clicked=n;break;}
    }
    if(clicked){focusNode=clicked.id;showNodeInfo(clicked);drawNetwork();}else{focusNode=null;document.getElementById('nwInfo').style.display='none';drawNetwork();}
  });
}
function resizeCanvas(){
  const c=document.getElementById('nwCanvas');
  const parent=c.parentElement;
  c.width=parent?parent.offsetWidth:820;
  c.height=600;
}
function toggleSidebar(){
  const side=document.getElementById('nwSide');
  const btn=document.getElementById('sbToggle');
  side.classList.toggle('collapsed');
  if(side.classList.contains('collapsed')){
    btn.textContent='▸ Show Panel';
  }else{
    btn.textContent='▸ Hide Panel';
  }
  setTimeout(()=>{resizeCanvas();drawNetwork();},50);
}
function filterNetwork(){drawNetwork();}
function clearFocus(){focusNode=null;document.getElementById('nwInfo').style.display='none';drawNetwork();}
function resetLayout(){scale=1;offsetX=0;offsetY=0;const c=document.getElementById('nwCanvas');
  const w=c.width||820;
  nwNodes.forEach((n,i)=>{nwPos[n.id]={x:w/2+Math.cos(i*0.5)*200,y:300+Math.sin(i*0.5)*200,vx:0,vy:0};});
  drawNetwork();}
function showNodeInfo(n){
  const a=NODE_ATTRIBUTES[n.id]||{};
  document.getElementById('nwInfo').style.display='block';
  document.getElementById('nwInfoContent').innerHTML=`
    <b>${n.label}</b> (${n.id})<br/>Betweenness: ${n.betweenness.toFixed(4)}<br/>
    Community: ${n.community}<br/>COG: ${a.cog_category||'-'}<br/>
    logFC: ${a.logFC!==null?a.logFC.toFixed(2):'-'}<br/>
    FDR: ${a.FDR!==null?a.FDR.toExponential(1):'-'}<br/>
    Essential: ${n.isEssential?'✓':'No'}`;
}
function getFilteredEdges(){
  const fm=document.getElementById('nwMethod').value;
  const fc=document.getElementById('nwClass').value;
  const q=document.getElementById('nwSearch').value.toLowerCase();
  return nwEdges.filter(e=>{
    if(fm!=='All methods'&&!e.method.includes(fm))return false;
    if(fc!=='All classes'&&e.class!==fc)return false;
    if(q){const la=NODE_ATTRIBUTES[e.source]?.gene_name||e.source,lb=NODE_ATTRIBUTES[e.target]?.gene_name||e.target;
      if(!(la+lb+e.source+e.target).toLowerCase().includes(q))return false;}
    return true;
  });
}
function drawNetwork(){
  const canvas=document.getElementById('nwCanvas');if(!canvas)return;
  resizeCanvas();
  const ctx=canvas.getContext('2d');
  ctx.clearRect(0,0,canvas.width,canvas.height);
  const edges=getFilteredEdges();
  const visibleNodes=new Set();
  edges.forEach(e=>{visibleNodes.add(e.source);visibleNodes.add(e.target);});
  if(focusNode){visibleNodes.add(focusNode);
    edges.forEach(e=>{if(e.source===focusNode)visibleNodes.add(e.target);if(e.target===focusNode)visibleNodes.add(e.source);});}
  const nodes=nwNodes.filter(n=>visibleNodes.has(n.id));
  for(let iter=0;iter<3;iter++){
    nodes.forEach(n=>{const p=nwPos[n.id];p.vx*=0.85;p.vy*=0.85;});
    for(let i=0;i<nodes.length;i++)for(let j=i+1;j<nodes.length;j++){
      const p1=nwPos[nodes[i].id],p2=nwPos[nodes[j].id];
      let dx=p2.x-p1.x,dy=p2.y-p1.y;let d=Math.sqrt(dx*dx+dy*dy)+0.1;
      const f=500/(d*d);dx/=d;dy/=d;
      p1.vx-=f*dx;p1.vy-=f*dy;p2.vx+=f*dx;p2.vy+=f*dy;
    }
    edges.forEach(e=>{
      if(!nwPos[e.source]||!nwPos[e.target])return;
      const p1=nwPos[e.source],p2=nwPos[e.target];
      let dx=p2.x-p1.x,dy=p2.y-p1.y;let d=Math.sqrt(dx*dx+dy*dy)+0.1;
      const f=0.01*d;dx/=d;dy/=d;
      p1.vx+=f*dx;p1.vy+=f*dy;p2.vx-=f*dx;p2.vy-=f*dy;
    });
    nodes.forEach(n=>{const p=nwPos[n.id];p.vx+=(canvas.width/2-p.x)*0.001;p.vy+=(canvas.height/2-p.y)*0.001;});
    nodes.forEach(n=>{const p=nwPos[n.id];p.x+=p.vx;p.y+=p.vy;
      p.x=Math.max(30,Math.min(canvas.width-30,p.x));p.y=Math.max(30,Math.min(canvas.height-30,p.y));});
  }
  ctx.save();ctx.translate(offsetX,offsetY);ctx.scale(scale,scale);
  const style=getComputedStyle(document.documentElement);
  const ecm={negative:style.getPropertyValue('--neg').trim(),positive:style.getPropertyValue('--pos').trim(),suppression:style.getPropertyValue('--sup').trim(),PPI:style.getPropertyValue('--ppi').trim()};
  edges.forEach(e=>{
    if(!nwPos[e.source]||!nwPos[e.target])return;
    const p1=nwPos[e.source],p2=nwPos[e.target];
    ctx.strokeStyle=ecm[e.class]||'#888';ctx.globalAlpha=focusNode&&e.source!==focusNode&&e.target!==focusNode?0.1:0.4;
    ctx.lineWidth=1;ctx.beginPath();ctx.moveTo(p1.x,p1.y);ctx.lineTo(p2.x,p2.y);ctx.stroke();
  });
  ctx.globalAlpha=1;
  const maxBetw=Math.max(...nodes.map(n=>n.betweenness||0));
  const palette=[];
  for(let i=0;i<''' + str(n_comm) + r''';i++)palette.push(`hsl(${i*360/''' + str(n_comm) + r'''},60%,55%)`);
  nodes.forEach(n=>{
    const p=nwPos[n.id];const rad=5+15*((n.betweenness||0)/maxBetw);
    ctx.fillStyle=palette[n.community%palette.length];
    ctx.globalAlpha=focusNode&&n.id!==focusNode?0.3:0.9;
    ctx.beginPath();ctx.arc(p.x,p.y,rad,0,2*Math.PI);ctx.fill();
    if(n.isEssential){ctx.strokeStyle='#ff0';ctx.lineWidth=2;ctx.stroke();}
    if(n.id===focusNode){ctx.strokeStyle='#fff';ctx.lineWidth=3;ctx.stroke();}
    if((n.betweenness||0)/maxBetw>0.3||n.id===focusNode){
      ctx.globalAlpha=1;ctx.fillStyle=style.getPropertyValue('--tx').trim();
      ctx.font='bold 9px Inter';ctx.textAlign='center';ctx.fillText(n.label,p.x,p.y-rad-3);
    }
  });
  ctx.globalAlpha=1;ctx.restore();
}

// ── CSV DOWNLOAD ──
function downloadCSV(){
  const headers=Object.keys(GI_DATABASE[0]||{});
  const csv=[headers.join(',')].concat(
    GI_DATABASE.map(e=>headers.map(h=>`"${String(e[h]||'').replace(/"/g,'""')}"`).join(','))
  ).join('\n');
  const blob=new Blob([csv],{type:'text/csv'});const url=URL.createObjectURL(blob);
  const a=document.createElement('a');a.href=url;a.download='pneumo_gi_database.csv';a.click();
  URL.revokeObjectURL(url);
}

init();
</script>
</body>
</html>'''


# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='S. pneumoniae Genetic Interaction Network Pipeline')
    parser.add_argument('--table-s3', required=True, help='Path to table_S3 xlsx')
    parser.add_argument('--table-s1', required=True, help='Path to table_S1 xlsx')
    parser.add_argument('--sup-data1', required=True, help='Path to Supplementary Data 1 xlsx')
    parser.add_argument('--sup-data6', required=True, help='Path to Supplementary Data 6 xlsx')
    parser.add_argument('--table-s4', required=True, help='Path to table_S4 xlsx')
    parser.add_argument('--output-dir', default='/mnt/results', help='Output directory')
    parser.add_argument('--z-threshold', type=float, default=4.0, help='|zStrains| threshold for S1')
    parser.add_argument('--string-score', type=float, default=0.7, help='STRING confidence threshold')
    parser.add_argument('--string-species', type=int, default=1313, help='STRING species ID')
    parser.add_argument('--de-fdr', type=float, default=0.05, help='FDR threshold for DE genes')
    parser.add_argument('--de-logfc', type=float, default=1.0, help='|logFC| threshold for DE genes')
    parser.add_argument('--top-hubs', type=int, default=20, help='Number of hub genes to report')
    parser.add_argument('--data-dir', default=os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data'),
                        help='Directory containing companion data files (tigr4_genes.json, orthology mappings, etc.)')
    args = parser.parse_args()

    config = Config(
        table_s3=args.table_s3, table_s1=args.table_s1,
        sup_data1=args.sup_data1, sup_data6=args.sup_data6,
        table_s4=args.table_s4, output_dir=args.output_dir,
        z_threshold=args.z_threshold, string_score=args.string_score,
        string_species=args.string_species, de_fdr=args.de_fdr,
        de_logfc=args.de_logfc, top_hubs=args.top_hubs,
        data_dir=args.data_dir,
    )

    # Create output directories
    for d in ['', 'tables', 'figures', 'network_graphml', 'website']:
        os.makedirs(os.path.join(config.output_dir, d), exist_ok=True)

    print("=" * 70)
    print("S. pneumoniae Genetic Interaction Network Pipeline")
    print("=" * 70)

    # ── Step 1: Data Loading ──
    print("\n[Step 1] Loading data...")
    rnaseq_df, locus_map = load_rnaseq_de(config.sup_data6)
    s3_df = load_dual_tnseq_s3(config.table_s3)
    s3_pairs = set()
    for _, row in s3_df.iterrows():
        s3_pairs.add(tuple(sorted([row['locusId1'], row['locusId2']])))
    s1_df = load_dual_tnseq_s1(config.table_s1, s3_pairs, config.z_threshold)
    supseq_df = load_supseq_hits(config.sup_data1, locus_map)
    fitness_df = load_rbtnseq_fitness(config.table_s4)

    # ── Step 2: STRING PPI ──
    print("\n[Step 2] Fetching STRING PPI...")
    gene_names = list(set(locus_map.values()))
    gene_names = [g for g in gene_names if g and g != '-']
    string_df = fetch_string_ppi(gene_names, config.string_species, config.string_score)

    # ── Step 3: Network Construction ──
    print("\n[Step 3] Building network...")
    G = build_network(s3_df, s1_df, supseq_df, string_df, rnaseq_df,
                      fitness_df, locus_map, config)

    # Save GraphML (sanitize None values — GraphML doesn't support NoneType)
    G_undirected = nx.Graph()
    for u, v, d in G.edges(data=True):
        clean_d = {k: ('' if v is None else v) for k, v in d.items()}
        G_undirected.add_edge(u, v, **clean_d)
    for n in G.nodes():
        nd = G.nodes[n]
        clean_nd = {k: ('' if v is None else v) for k, v in nd.items()}
        G_undirected.add_node(n, **clean_nd)
    graphml_path = os.path.join(config.output_dir, 'network_graphml', 'network.graphml')
    nx.write_graphml(G_undirected, graphml_path)
    print(f"  Saved: {graphml_path}")

    # ── Step 4: Topology Analysis ──
    print("\n[Step 4] Computing network topology...")
    centrality_df, Gu = compute_centrality(G)
    hub_path = os.path.join(config.output_dir, 'tables', 'hub_genes.csv')
    centrality_df.head(config.top_hubs).to_csv(hub_path, index=False)
    print(f"  Saved: {hub_path}")

    # ── Step 5: Community Detection ──
    print("\n[Step 5] Detecting communities...")
    comm_map, modularity = detect_communities(Gu)

    # ── Step 6: Perturbed Pathways ──
    print("\n[Step 6] Identifying perturbed pathways...")
    de_enrichment = test_de_enrichment(comm_map, G, config.de_fdr)
    cog_chars = characterize_cog(comm_map, G)

    # Merge and save
    community_df = de_enrichment.merge(cog_chars, on='community', how='left')
    comm_path = os.path.join(config.output_dir, 'tables', 'community_annotations.csv')
    community_df.to_csv(comm_path, index=False)
    print(f"  Saved: {comm_path}")

    # Save combined edge table
    edge_rows = []
    for u, v, d in G.edges(data=True):
        edge_rows.append({
            'locusId1': u, 'locusId2': v,
            'gene1': G.nodes[u].get('gene_name', u),
            'gene2': G.nodes[v].get('gene_name', v),
            **d
        })
    edge_df = pd.DataFrame(edge_rows)
    edge_path = os.path.join(config.output_dir, 'tables', 'genetic_interactions_combined.csv')
    edge_df.to_csv(edge_path, index=False)
    print(f"  Saved: {edge_path}")

    # ── Step 7: Visualization ──
    print("\n[Step 7] Creating visualizations...")
    create_centrality_plots(centrality_df, config.output_dir, config.top_hubs)
    create_community_plot(G, Gu, comm_map, centrality_df, config.output_dir)
    create_de_enrichment_heatmap(de_enrichment, config.output_dir)
    create_edge_composition_plot(G, comm_map, config.output_dir)
    create_pyvis_html(G, comm_map, centrality_df, config.output_dir)

    # ── Step 8: Webpage ──
    print("\n[Step 8] Generating community webpage...")
    generate_webpage(G, centrality_df, comm_map, de_enrichment, cog_chars, config)

    # ── Copy script to output ──
    import shutil
    script_src = os.path.abspath(__file__)
    script_dst = os.path.join(config.output_dir, 'grn_pipeline.py')
    shutil.copy(script_src, script_dst)  # copy (not copy2) — S3 mount doesn't support utime
    print(f"\n  Script saved: {script_dst}")

    print("\n" + "=" * 70)
    print("Pipeline complete!")
    print(f"  Nodes: {G.number_of_nodes()}, Edges: {G.number_of_edges()}")
    print(f"  Communities: {len(set(comm_map.values()))}, Modularity: {modularity:.4f}")
    print(f"  Top hub: {centrality_df.iloc[0]['gene_name']} (betweenness={centrality_df.iloc[0]['betweenness']:.4f})")
    print(f"  Outputs in: {config.output_dir}")
    print("=" * 70)


if __name__ == '__main__':
    main()
